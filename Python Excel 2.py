from openpyxl import Workbook # pip install openpyxl 
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter
 # see example 1 or
listA=['DistanceA (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
listB=['DistanceB (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
L=[listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB]
wb = Workbook()
ws1 = wb.active # work sheet
ws1.title = "Pyxl"
ft_1 = Font(name='Calibri',color=colors.BLUE, bold=True, size=11,underline='none')
ft_2 = Font(name='Calibri',color=colors.GREEN, bold=True, size=11,underline='none')
ft_3 = Font(name='Calibri',color=colors.RED, bold=True, size=11,underline='none')
ft_4 = Font(name='Calibri',color=colors.BLACK, bold=True, size=11,underline='none')

for i in range(len(listA)):
    for j in range (len(L)):
        _ = ws1.cell(column=j+2, row=i+2, value=L[j][i])
        if i==len(listA)-1 :
            _ = ws1.cell(column=j+2, row=i+3, value='=SUM('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')') 
            _.font=ft_1
            _ = ws1.cell(column=j+2, row=i+4, value='=average('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_2
            _.number_format='#,#0.0'
            _ = ws1.cell(column=j+2, row=i+5, value='=max('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_3
    # ws1.append([listA[row]])

for i, name in enumerate(['Summation', 'Average', 'Maximum']):
    _ = ws1.cell(column=1, row=len(listA)+2+i, value=name)
    _.font=ft_4 



# ft_h = Font(name='Calibri',color=colors.BLUE, bold=True, size=11,underline='double')
# ft_b = Font(name='Calibri',color=colors.RED, bold=True, size=11,underline='single')
# a1 = ws1['A1']
# b1 = ws1['B1']
# a1.font = ft_h
# b1.font = ft_h

# for i in range (2,13):
#         a=ws1['A'+str(i)]
#         b=ws1['B'+str(i)]
#         a.font=ft_b
#         b.font=ft_b

# ws1["A14"]='Summation='
# ws1["B14"]='=SUM(B3:B13)'
# ws1["C14"]='=SUM(C3:C13)'


# ws1["A15"]='Maximum='
# ws1["B15"]='=max(B3:B13)'
# ws1["C15"]='=max(C3:C13)'

# # Average
# ws1["A16"]='Average='
# ws1["C16"]='=average(C3:C13)'
# ws1["B16"]='=average(B3:B13)'

wb.save('xlfile.xlsx')