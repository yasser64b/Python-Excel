from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, fills
import numpy as np

#  loading the file 
wb = load_workbook('test results.xlsx')
SheetNames= wb.sheetnames
MeanMax=[]
cols='abcdefghijklmnopqrstuvwxyz'

for names in SheetNames:
    sheet_ranges = wb[names]
    for col in cols:
        data=[]
        header=''
        for row in range (1, 200):
            cellValue=sheet_ranges[col+str(row)].value
            if type(cellValue) is str:
                header=cellValue
            elif type(cellValue) is int or type(cellValue) is float:
                data.append(cellValue)
        if len(data)>0:
            MeanMax.append([names, header, len(data), np.mean(data), np.max(data)])
# print(MeanMax)

wb1 = Workbook()
ws1 = wb1.active # work sheet
ws1.title = "tests"

for j in range (0,21):
    means=[]
    maxes=[]
    for i in range (len(MeanMax)):
        if MeanMax[i][1]=='Test'+str(j):
            means.append(MeanMax[i][3])
            maxes.append(MeanMax[i][4])
    if len(means)>0:
        MeanT=np.mean(means)
        MaxT=np.max(maxes)
        _ = ws1.cell(column=j+1, row=1, value='test'+str(j))
        _ = ws1.cell(column=1, row=1, value='test ranges')
        _ = ws1.cell(column=j+1, row=2, value=MeanT)
        _ = ws1.cell(column=1, row=2, value='Average')
        _ = ws1.cell(column=j+1, row=3, value=MaxT)
        _ = ws1.cell(column=1, row=3, value='Maximum')  

thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                right=Side(border_style='dashed',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
                
thick_border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='medium',color='FF000000')
                )
Double_border = Border(left=Side(border_style='dashed',color='FF000000'),
                right=Side(border_style='dashed',color='FF000000'),
                top=Side(border_style='double',color='FF000000'),
                bottom=Side(border_style='double',color='FF000000')
                )

#Define fill formating 
fill_cell = PatternFill(fill_type=fills.FILL_SOLID,
                        start_color='00FFFF00',end_color='00FFFF00') 
            
for k in range (1,4):
    for l in range (1,22):
        ws1.cell(k,l).number_format='#,#0.0'
        ws1.cell(k,l).border=thin_border
        if k==1:
            ws1.cell(k,l).fill=fill_cell

wb1.save('analyzed data.xlsx')