# See full Toturial at my Youtube Channel(YB TV): https://www.youtube.com/channel/UCvnhhDKv5takEN412dmVW8g/featured
# GitHab Page:https://github.com/yasser64b/
#Email: big3del@gmail.com 

from openpyxl import Workbook # pip install openpyxl 
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series, LineChart, ScatterChart
 # see example 1 or
listA=['DistanceA (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
listB=['DistanceB (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
L=[listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB]
wb = Workbook()
ws1 = wb.active # work sheet
ws1.title = "Pyxl"
ft_1 = Font(name='Calibri',color=colors.BLUE, bold=True, size=11,underline='none')
ft_2 = Font(name='Calibri',color=colors.GREEN, bold=True, size=11,underline='none')
ft_3 = Font(name='Calibri',color=colors.RED, bold=True, size=11,underline='none')
ft_4 = Font(name='Calibri',color=colors.BLACK, bold=True, size=11,underline='none')

for i in range(len(listA)):
    for j in range (len(L)):
        _ = ws1.cell(column=j+2, row=i+2, value=L[j][i])
        if i==len(listA)-1 :
            _ = ws1.cell(column=j+2, row=i+3, value='=SUM('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')') 
            _.font=ft_1
            _ = ws1.cell(column=j+2, row=i+4, value='=average('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_2
            _.number_format='#,#0.0'
            _ = ws1.cell(column=j+2, row=i+5, value='=max('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_3
    # ws1.append([listA[row]])

for i, name in enumerate(['Summation', 'Average', 'Maximum']):
    _ = ws1.cell(column=1, row=len(listA)+2+i, value=name)
    _.font=ft_4 


# drw graphs
for j in range (len(L)):
    values = Reference(ws1, min_col=j+2, min_row=2, max_col=j+2, max_row=1+len(listA))
    chart = LineChart()
    chart.title ='Chart-'+str(j+1)
    chart.y_axis.title = 'Size'
    chart.x_axis.title = 'Number-'+str(j+1)
    chart.add_data(values) 
    if j<int(len(L)/2):
        ws1.add_chart(chart, get_column_letter(20)+str((j+2)+j*14))
    else:
        ws1.add_chart(chart, get_column_letter(30)+str((j-int(len(L)/2)+2)+(j-int(len(L)/2))*14))


wb.save('xlfile.xlsx')
