# See full Toturial at my Youtube Channel(YB TV): https://www.youtube.com/channel/UCvnhhDKv5takEN412dmVW8g/featured
# GitHab Page:https://github.com/yasser64b/
#Email: big3del@gmail.com


from openpyxl import Workbook #pip install openpyxl  , See the link below
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills

wb=Workbook()
ws1=wb.active
ws1.title="PyXL"
ws1.sheet_properties.tabColor = '000000FF'

# define border formats 

thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                right=Side(border_style='dashed',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
                
thick_border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='medium',color='FF000000')
                )
Double_border = Border(left=Side(border_style='dashed',color='FF000000'),
                right=Side(border_style='dashed',color='FF000000'),
                top=Side(border_style='double',color='FF000000'),
                bottom=Side(border_style='double',color='FF000000')
                )

#Define fill formating 
fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')

#define size of the table 
row_num=5
col_num=6 
#location of the Table 
row_loc=3
col_loc=3

#Number of Tables 
Table_num=10
dis=2 # distance between the tables 
for _ in range(Table_num):
    for i in range (row_loc,row_loc+row_num):
            for j in range (col_loc,col_num+col_loc):
                ws1.cell(row=i+1, column=j+1).border=thin_border
                if i==row_loc:
                    ws1.cell(row=i+1, column=j+1).border=Double_border
                    ws1.cell(row=i+1, column=j+1).fill=fill_cell
                if i==row_loc+row_num-1:
                    ws1.cell(row=i+1, column=j+1).border=thick_border
    row_loc= row_loc+row_num+dis 
wb.save('xlpy4.xlsx')


